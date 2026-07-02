
import os
import json
import io
import boto3
from aws_lambda_powertools import Logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = Logger(service="excel-generator")
s3_client = boto3.client("s3", region_name="us-east-1")


def nome_amigavel(chave):
    return str(chave).replace("_", " ").replace("$", "USD").title()


def valor_para_celula(valor):
    if valor is None:
        return ""
    if isinstance(valor, (dict, list)):
        return json.dumps(valor, ensure_ascii=False)
    return str(valor)


def aplicar_estilo_corporativo(ws):
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = "A2"

    azul_marinho_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    azul_secao_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    cinza_claro_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    fonte_cabecalho = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    fonte_secao = Font(name="Arial", size=11, bold=True, color="1E3A8A")
    fonte_comum = Font(name="Arial", size=10, bold=False, color="334155")
    fonte_negrito = Font(name="Arial", size=10, bold=True, color="0F172A")

    alinhamento_esquerda = Alignment(horizontal="left", vertical="top", wrap_text=True)
    alinhamento_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

    borda_fina = Side(border_style="thin", color="CBD5E1")
    caixa_borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

    max_col = max(ws.max_column, 2)

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 22

        primeira_celula = ws.cell(row=row_idx, column=1)
        eh_secao = isinstance(primeira_celula.value, str) and primeira_celula.value.startswith("## ")

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = caixa_borda
            cell.alignment = alinhamento_esquerda

            if row_idx == 1:
                cell.fill = azul_marinho_fill
                cell.font = fonte_cabecalho
                cell.alignment = alinhamento_centro
            elif eh_secao:
                cell.fill = azul_secao_fill
                cell.font = fonte_secao
            elif row_idx % 2 == 0:
                cell.fill = cinza_claro_fill
                cell.font = fonte_negrito if col_idx == 1 else fonte_comum
            else:
                cell.font = fonte_negrito if col_idx == 1 else fonte_comum


def auto_ajustar_largura_colunas(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        if col_letter == "A":
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 22), 38)
        else:
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 15), 35)


def adicionar_secao(ws, titulo):
    ws.append([f"## {titulo}"])


def adicionar_linha_simples(ws, chave, valor):
    ws.append([nome_amigavel(chave), valor_para_celula(valor)])


def renderizar_dict(ws, dados, titulo=None):
    if titulo:
        adicionar_secao(ws, nome_amigavel(titulo))

    for chave, valor in dados.items():
        renderizar_campo(ws, chave, valor)


def renderizar_lista_de_dicts(ws, titulo, lista):
    adicionar_secao(ws, nome_amigavel(titulo))

    if not lista:
        ws.append(["Sem dados identificados"])
        return

    cabecalhos = []
    for item in lista:
        if isinstance(item, dict):
            for chave in item.keys():
                if chave not in cabecalhos:
                    cabecalhos.append(chave)

    if not cabecalhos:
        for idx, item in enumerate(lista, start=1):
            ws.append([f"Item {idx}", valor_para_celula(item)])
        return

    ws.append([nome_amigavel(c) for c in cabecalhos])

    for item in lista:
        if isinstance(item, dict):
            ws.append([valor_para_celula(item.get(c)) for c in cabecalhos])
        else:
            ws.append([valor_para_celula(item)])


def renderizar_lista(ws, titulo, lista):
    if all(isinstance(item, dict) for item in lista):
        renderizar_lista_de_dicts(ws, titulo, lista)
    else:
        adicionar_secao(ws, nome_amigavel(titulo))
        for idx, item in enumerate(lista, start=1):
            ws.append([f"Item {idx}", valor_para_celula(item)])


def renderizar_campo(ws, chave, valor):
    if isinstance(valor, dict):
        adicionar_secao(ws, nome_amigavel(chave))

        for sub_chave, sub_valor in valor.items():
            if isinstance(sub_valor, list):
                renderizar_lista(ws, sub_chave, sub_valor)
            elif isinstance(sub_valor, dict):
                renderizar_dict(ws, sub_valor, sub_chave)
            else:
                adicionar_linha_simples(ws, sub_chave, sub_valor)

    elif isinstance(valor, list):
        renderizar_lista(ws, chave, valor)

    else:
        adicionar_linha_simples(ws, chave, valor)


def handler(event, context):
    try:
        logger.info(f"Iniciando engine de renderização de planilhas para o evento: {json.dumps(event)}")

        package_id = event.get("package_id")
        s3_key_json = event.get("s3_key_resultado")
        bucket = event.get("bda_output_bucket") or os.environ.get("BUCKET_SAIDA", "credifacil-docs-saida-dev")
        arquivo_original = event.get("arquivo_original", "documento_analisado.pdf")

        if not s3_key_json or not package_id:
            raise ValueError("Propriedades 's3_key_resultado' ou 'package_id' ausentes no payload.")

        s3_response = s3_client.get_object(Bucket=bucket, Key=s3_key_json)
        payload_dados = json.loads(s3_response["Body"].read().decode("utf-8"))

        wb = Workbook()
        ws = wb.active
        ws.title = "Metadados Estruturados"

        ws.append(["Propriedade Analisada", "Valor Identificado"])

        if "cliente" in payload_dados and "validacao" in payload_dados:
            cliente = payload_dados["cliente"]
            validacao = payload_dados["validacao"]

            ws.append(["Nome Completo do Proponente", cliente.get("nome", "Não Identificado")])
            ws.append(["Documento de Identificação", cliente.get("documento_identificacao", "Não Informado")])
            ws.append(["Score de Crédito Atribuído", f"{cliente.get('score_credito', {}).get('valor', 0)} Pontos"])
            ws.append(["Classificação de Risco", str(cliente.get("classificacao_risco", {}).get("categoria", "INCONCLUSIVO")).upper()])
            ws.append(["Parecer / Justificativa Técnica", cliente.get("classificacao_risco", {}).get("justificativa", "")])

            for chk_chave, chk_val in validacao.items():
                nome_chk = nome_amigavel(chk_chave)
                status_txt = "✅ CONSISTENTE / PRESENTE" if chk_val is True else "❌ DIVERGENTE / AUSENTE" if chk_val is False else "⚪ NÃO AVALIADO"
                ws.append([f"Checklist: {nome_chk}", status_txt])

        else:
            campos_extraidos = payload_dados.get("dados_extraidos_do_documento", {})

            for chave, valor in campos_extraidos.items():
                renderizar_campo(ws, chave, valor)

        aplicar_estilo_corporativo(ws)
        auto_ajustar_largura_colunas(ws)

        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        if s3_key_json and "customer_consolidated" in s3_key_json:
            nome_limpo_arquivo = "customer_consolidated"
        else:
            nome_limpo_arquivo = (
                arquivo_original
                .replace(".pdf", "")
                .replace(".png", "")
                .replace(".jpg", "")
                .replace(".jpeg", "")
            )

        s3_target_excel_key = f"results/planilhas/{package_id}/excel_metadados_{nome_limpo_arquivo}.xlsx"

        logger.info(f"Salvando planilha no S3: {s3_target_excel_key}")

        s3_client.put_object(
            Bucket=bucket,
            Key=s3_target_excel_key,
            Body=output_buffer.getvalue(),
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        return {
            "status": "SUCCESS",
            "package_id": package_id,
            "excel_s3_key": s3_target_excel_key
        }

    except Exception as e:
        logger.error(f"Falha ao gerar planilha executiva no backend: {str(e)}")
        raise e
